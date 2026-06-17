import csv
import json
import re
from pathlib import Path

import frappe
from frappe.utils import now


SYNC_FIELDS = ["customer", "client_sku", "product_name", "product_description", "uom", "barcode", "product_image", "status"]
IMPORT_COLUMNS = ["client_sku", "product_name", "product_description", "uom", "barcode", "product_image", "status", "notes"]
REQUIRED_IMPORT_COLUMNS = {"client_sku", "product_name"}


def clean_code(value):
    cleaned = re.sub(r"[^A-Z0-9]+", "-", str(value or "").upper()).strip("-")
    return cleaned or "ITEM"


def customer_prefix(customer):
    words = [word for word in re.split(r"[^A-Za-z0-9]+", customer or "") if word.lower() not in {"demo", "client", "customer"}]
    return clean_code(words[-1] if words else customer)[:12]


def product_snapshot(product):
    return {field: product.get(field) for field in SYNC_FIELDS}


def json_dumps(data):
    return json.dumps(data or {}, sort_keys=True, ensure_ascii=False, indent=2)


def find_existing_item(product):
    if product.item_code and frappe.db.exists("Item", product.item_code):
        return product.item_code

    return frappe.db.get_value(
        "Item",
        {
            "owner_client": product.customer,
            "client_sku": product.client_sku,
        },
        "name",
    )


def generated_item_code(product):
    base = f"{customer_prefix(product.customer)}-{clean_code(product.client_sku)}"
    existing = frappe.db.get_value(
        "Item",
        {
            "owner_client": product.customer,
            "client_sku": product.client_sku,
        },
        "name",
    )
    if existing:
        return existing

    if not frappe.db.exists("Item", base):
        return base

    counter = 2
    while frappe.db.exists("Item", f"{base}-{counter}"):
        counter += 1
    return f"{base}-{counter}"


def set_if_field(doc, fieldname, value):
    if doc.meta.has_field(fieldname):
        setattr(doc, fieldname, value)


def ensure_barcode(item, barcode):
    if not barcode:
        return
    if not any(row.barcode == barcode for row in item.get("barcodes", [])):
        item.append("barcodes", {"barcode": barcode})


def change_action(old_snapshot, new_snapshot):
    if not old_snapshot:
        return "Deactivated" if new_snapshot.get("status") == "Inactive" else "Created"
    if old_snapshot.get("status") != "Inactive" and new_snapshot.get("status") == "Inactive":
        return "Deactivated"
    return "Updated"


def insert_log(product, action, old_snapshot, new_snapshot, notes=None):
    log = frappe.new_doc("Three PL Client Product Change Log")
    log.product = product.name
    log.customer = product.customer
    log.item_code = product.item_code
    log.action = action
    log.changed_by = product.modified_by or product.owner
    log.change_datetime = product.modified or now()
    log.old_values = json_dumps(old_snapshot)
    log.new_values = json_dumps(new_snapshot)
    log.notes = notes
    log.insert(ignore_permissions=True)
    return log.name


def normalize_status(value):
    value = str(value or "Active").strip() or "Active"
    if value.lower() in {"active", "enabled", "1", "yes", "y"}:
        return "Active"
    if value.lower() in {"inactive", "disabled", "0", "no", "n"}:
        return "Inactive"
    raise RuntimeError(f"Unsupported product status: {value}")


def sync_product(product_name):
    product = frappe.get_doc("Three PL Client Product", product_name)
    snapshot = product_snapshot(product)
    old_snapshot = json.loads(product.last_synced_snapshot or "{}")

    try:
        item_code = find_existing_item(product) or generated_item_code(product)
        item = frappe.get_doc("Item", item_code) if frappe.db.exists("Item", item_code) else frappe.new_doc("Item")

        if item.is_new():
            item.item_code = item_code
            item.item_group = "Products"
            item.is_stock_item = 1

        item.item_name = product.product_name
        item.description = product.product_description or product.product_name
        item.item_group = item.item_group or "Products"
        item.stock_uom = product.uom or "Nos"
        item.disabled = 1 if product.status == "Inactive" else 0
        set_if_field(item, "owner_client", product.customer)
        set_if_field(item, "client_sku", product.client_sku)
        set_if_field(item, "client_product_name", product.product_name)
        set_if_field(item, "image", product.product_image)
        ensure_barcode(item, product.barcode)
        item.save(ignore_permissions=True)

        product.item_code = item.name
        product.sync_status = "Synced"
        product.sync_error = ""
        product.last_synced_at = now()
        product.last_synced_snapshot = json_dumps(snapshot)
        product.save(ignore_permissions=True)

        if old_snapshot != snapshot:
            insert_log(product, change_action(old_snapshot, snapshot), old_snapshot, snapshot, "Product card synchronized to ERPNext Item.")
        else:
            insert_log(product, "Synced", old_snapshot, snapshot, "Product card re-synchronized without client field changes.")

        return item.name
    except Exception as exc:
        product.sync_status = "Failed"
        product.sync_error = str(exc)
        product.save(ignore_permissions=True)
        insert_log(product, "Sync Failed", old_snapshot, snapshot, str(exc))
        raise


def sync_client_products():
    synced = []
    failed = []
    for product_name in frappe.get_all("Three PL Client Product", pluck="name"):
        product = frappe.get_doc("Three PL Client Product", product_name)
        current_snapshot = product_snapshot(product)
        previous_snapshot = json.loads(product.last_synced_snapshot or "{}")
        if product.sync_status == "Synced" and current_snapshot == previous_snapshot:
            continue
        try:
            synced.append(sync_product(product_name))
            frappe.db.commit()
        except Exception as exc:
            failed.append((product_name, str(exc)))
            frappe.db.commit()
    return synced, failed


def file_path(file_url):
    file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not file_name:
        raise RuntimeError(f"Import file not found: {file_url}")
    file_doc = frappe.get_doc("File", file_name)
    return Path(file_doc.get_full_path())


def parse_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def parse_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX import requires openpyxl in the ERPNext backend. Upload CSV instead.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    parsed = []
    for row in rows[1:]:
        parsed.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return parsed


def import_rows(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return parse_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(path)
    raise RuntimeError("Unsupported import format. Upload .csv or .xlsx.")


def validate_import_headers(rows):
    if not rows:
        raise RuntimeError("Import file has no product rows.")
    headers = set(rows[0])
    missing = sorted(REQUIRED_IMPORT_COLUMNS - headers)
    if missing:
        raise RuntimeError("Import file misses required columns: " + ", ".join(missing))


def upsert_product_from_row(import_doc, row, row_number):
    client_sku = str(row.get("client_sku") or "").strip()
    product_name = str(row.get("product_name") or "").strip()
    if not client_sku:
        raise RuntimeError(f"Row {row_number}: client_sku is required")
    if not product_name:
        raise RuntimeError(f"Row {row_number}: product_name is required")

    existing = frappe.db.get_value(
        "Three PL Client Product",
        {"customer": import_doc.customer, "client_sku": client_sku},
        "name",
    )
    product = frappe.get_doc("Three PL Client Product", existing) if existing else frappe.new_doc("Three PL Client Product")
    product.customer = import_doc.customer
    product.client_sku = client_sku
    product.product_name = product_name
    product.product_description = str(row.get("product_description") or "").strip()
    product.uom = str(row.get("uom") or "Nos").strip() or "Nos"
    product.barcode = str(row.get("barcode") or "").strip()
    product.product_image = str(row.get("product_image") or "").strip()
    product.status = normalize_status(row.get("status"))
    product.notes = str(row.get("notes") or "").strip()
    product.sync_status = "Pending"
    product.save(ignore_permissions=True)
    if product.owner != import_doc.owner:
        frappe.db.set_value(product.doctype, product.name, "owner", import_doc.owner, update_modified=False)
        product.owner = import_doc.owner
    return product.name


def process_product_import(import_name):
    import_doc = frappe.get_doc("Three PL Client Product Import", import_name)
    rows = import_rows(file_path(import_doc.import_file))
    validate_import_headers(rows)

    errors = []
    applied = []
    for index, row in enumerate(rows, start=2):
        if not any(row.values()):
            continue
        try:
            applied.append(upsert_product_from_row(import_doc, row, index))
        except Exception as exc:
            errors.append(str(exc))

    import_doc.rows_total = len([row for row in rows if any(row.values())])
    import_doc.rows_applied = len(applied)
    import_doc.processed_at = now()
    import_doc.error_log = "\n".join(errors)
    import_doc.status = "Failed" if errors else "Applied"
    import_doc.save(ignore_permissions=True)

    for product_name in applied:
        sync_product(product_name)
    return import_doc.name


def sync_product_imports():
    applied = []
    failed = []
    for import_name in frappe.get_all("Three PL Client Product Import", filters={"status": "Pending"}, pluck="name"):
        try:
            applied.append(process_product_import(import_name))
            frappe.db.commit()
        except Exception as exc:
            import_doc = frappe.get_doc("Three PL Client Product Import", import_name)
            import_doc.status = "Failed"
            import_doc.processed_at = now()
            import_doc.error_log = str(exc)
            import_doc.save(ignore_permissions=True)
            failed.append((import_name, str(exc)))
            frappe.db.commit()
    return applied, failed


def main():
    imports_applied, imports_failed = sync_product_imports()
    synced, failed = sync_client_products()
    frappe.db.commit()
    print(f"Applied client product imports: {len(imports_applied)}")
    for import_name in imports_applied:
        print(import_name)
    print(f"Failed client product imports: {len(imports_failed)}")
    for import_name, error in imports_failed:
        print(f"{import_name}: {error}")
    print(f"Synced client products: {len(synced)}")
    for item_code in synced:
        print(item_code)
    print(f"Failed client product syncs: {len(failed)}")
    for product_name, error in failed:
        print(f"{product_name}: {error}")


if __name__ == "__main__":
    main()
